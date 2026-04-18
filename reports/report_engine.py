"""
Report Engine - Gerador de Relatório Excel
==========================================
Gera relatório Excel estruturado com quatro abas:
    1. Resumo Executivo
    2. Resultados
    3. Plano de Ação
    4. Comparação (quando disponível)
"""

from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Paleta de cores ──────────────────────────────────────────────────────────
COR_CABECALHO    = "1F3864"
COR_CONFORME     = "E2EFDA"
COR_NAO_CONFORME = "FFDAD6"
COR_ATENCAO      = "FCE4D6"
COR_PENDENTE     = "F2F2F2"
COR_SCORE = {
    0: "FFDAD6",
    1: "FCE4D6",
    2: "D5E8F0",
    3: "E2EFDA",
}

STATUS_CORES = {
    "CONFORME":     COR_CONFORME,
    "NÃO CONFORME": COR_NAO_CONFORME,
    "ATENÇÃO":      COR_ATENCAO,
}


# ── Helpers de estilo ────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Arial")


def _alinhar(horizontal="left", vertical="center"):
    return Alignment(horizontal=horizontal, vertical=vertical,
                     wrap_text=True)


def _borda():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _celula(ws, linha, col, valor, fill=None, bold=False, align="left", size=11):
    cel = ws.cell(row=linha, column=col, value=valor)
    if fill:
        cel.fill = _fill(fill)
    cel.font = _font(bold=bold, size=size)
    cel.alignment = _alinhar(align)
    cel.border = _borda()
    return cel


def _cabecalho(ws, linha, colunas, larguras):
    for col, (texto, largura) in enumerate(zip(colunas, larguras), 1):
        cel = ws.cell(row=linha, column=col, value=texto)
        cel.fill = _fill(COR_CABECALHO)
        cel.font = _font(bold=True, color="FFFFFF")
        cel.alignment = _alinhar("center")
        cel.border = _borda()
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = largura


# ── Aba 1: Resumo Executivo ──────────────────────────────────────────────────
def _aba_resumo(wb, scores, score_geral, timestamp):
    ws = wb.active
    ws.title = "Resumo Executivo"

    # Título
    ws.merge_cells("A1:F1")
    cel = ws["A1"]
    cel.value = "BlindSpot - Relatório de Auditoria de Segurança"
    cel.fill = _fill(COR_CABECALHO)
    cel.font = _font(bold=True, color="FFFFFF", size=14)
    cel.alignment = _alinhar("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    cel = ws["A2"]
    cel.value = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cel.fill = _fill("D5E8F0")
    cel.font = _font(size=10)
    cel.alignment = _alinhar("center")

    # Score geral
    ws.merge_cells("A4:F4")
    cel = ws["A4"]
    cel.value = f"Score Geral: {score_geral['score_geral']} / 3.0  -  {score_geral['nivel_geral']}"
    cel.fill = _fill(COR_SCORE.get(round(score_geral['score_geral']), "FCE4D6"))
    cel.font = _font(bold=True, size=13)
    cel.alignment = _alinhar("center")
    ws.row_dimensions[4].height = 25

    # Cabeçalho da tabela
    colunas  = ["Módulo", "Verificações", "Conformes", "Não Conformes", "Conformidade (%)", "Maturidade"]
    larguras = [18, 15, 12, 15, 18, 20]
    _cabecalho(ws, 6, colunas, larguras)

    linha = 7
    for modulo, dados in scores.items():
        score = dados.get("score", 0)
        _celula(ws, linha, 1, modulo, bold=True)
        _celula(ws, linha, 2, dados.get("total", 0), align="center")
        _celula(ws, linha, 3, dados.get("conformes", 0), align="center")
        _celula(ws, linha, 4, dados.get("nao_conformes", 0), align="center")
        _celula(ws, linha, 5, f"{dados.get('percentual', 0)}%", align="center")
        _celula(ws, linha, 6, f"{score} — {dados.get('nivel', '')}", fill=COR_SCORE.get(score), bold=True)
        ws.row_dimensions[linha].height = 18
        linha += 1


# ── Aba 2: Resultados ────────────────────────────────────────────────────────
def _aba_resultados(wb, resultados):
    ws = wb.create_sheet("Resultados")

    colunas  = ["Módulo", "Controle ISO/LGPD", "Função NIST CSF", "Descrição", "Status", "Evidência", "Recomendação"]
    larguras = [14, 22, 22, 38, 16, 35, 38]
    _cabecalho(ws, 1, colunas, larguras)

    linha = 2
    for modulo, verificacoes in resultados.items():
        for v in verificacoes:
            status = v.get("status", "")
            fill   = STATUS_CORES.get(status, COR_PENDENTE)
            _celula(ws, linha, 1, v.get("modulo", modulo), bold=True)
            _celula(ws, linha, 2, v.get("controle_iso", ""))
            _celula(ws, linha, 3, v.get("funcao_nist", ""))
            _celula(ws, linha, 4, v.get("descricao", ""))
            _celula(ws, linha, 5, status, fill=fill, bold=True, align="center")
            _celula(ws, linha, 6, v.get("evidencia", ""))
            _celula(ws, linha, 7, v.get("remediacao", ""))
            ws.row_dimensions[linha].height = 40
            linha += 1


# ── Aba 3: Plano de Ação ────────────────────────────────────────────────────
def _aba_plano(wb, resultados):
    ws = wb.create_sheet("Plano de Ação")

    colunas  = ["Módulo", "Controle ISO/LGPD", "Recomendação", "Prioridade", "Função NIST CSF", "Responsável", "Prazo"]
    larguras = [14, 22, 40, 12, 22, 22, 15]
    _cabecalho(ws, 1, colunas, larguras)

    linha = 2
    for modulo, verificacoes in resultados.items():
        nao_conformes = [v for v in verificacoes if v.get("status") == "NÃO CONFORME"]
        for v in nao_conformes:
            _celula(ws, linha, 1, v.get("modulo", modulo), bold=True)
            _celula(ws, linha, 2, v.get("controle_iso", ""))
            _celula(ws, linha, 3, v.get("remediacao", ""), fill=COR_NAO_CONFORME)
            _celula(ws, linha, 4, "Alta", fill=COR_NAO_CONFORME, bold=True, align="center")
            _celula(ws, linha, 5, v.get("funcao_nist", ""))
            _celula(ws, linha, 6, "")
            _celula(ws, linha, 7, "")
            ws.row_dimensions[linha].height = 40
            linha += 1

    if linha == 2:
        ws.merge_cells("A2:G2")
        cel = ws["A2"]
        cel.value = "Nenhuma não conformidade identificada."
        cel.fill = _fill(COR_CONFORME)
        cel.font = _font(bold=True, color="375623")
        cel.alignment = _alinhar("center")


# ── Aba 4: Comparação ───────────────────────────────────────────────────────
def _aba_comparacao(wb, delta):
    if not delta:
        return

    ws = wb.create_sheet("Comparação")

    colunas  = ["Módulo", "Score Anterior", "Score Atual", "Delta", "Tendência", "Data Anterior"]
    larguras = [16, 16, 14, 10, 20, 22]
    _cabecalho(ws, 1, colunas, larguras)

    linha = 2
    for modulo, dados in delta.items():
        d    = dados["delta"]
        fill = "E2EFDA" if d > 0 else ("FFDAD6" if d < 0 else "F2F2F2")

        # Remove códigos ANSI da tendência para o Excel
        tendencia = dados["tendencia"]
        for codigo in ["\033[32m", "\033[31m", "\033[90m", "\033[0m"]:
            tendencia = tendencia.replace(codigo, "")

        _celula(ws, linha, 1, modulo, bold=True)
        _celula(ws, linha, 2, dados["score_anterior"], align="center")
        _celula(ws, linha, 3, dados["score_atual"], align="center")
        _celula(ws, linha, 4, dados["delta_texto"], fill=fill, bold=True, align="center")
        _celula(ws, linha, 5, tendencia, fill=fill)
        _celula(ws, linha, 6, dados.get("data_anterior", ""))
        ws.row_dimensions[linha].height = 18
        linha += 1


# ── Entry point ──────────────────────────────────────────────────────────────
def gerar_relatorio(resultados, scores, score_geral, delta, output_file):
    """
    Gera o relatório Excel completo.

    Args:
        resultados:  dict {modulo: [verificacoes]}
        scores:      dict {modulo: resultado do calcular_score}
        score_geral: dict com score_geral e nivel_geral
        delta:       dict de comparação ou None
        output_file: caminho de saída
    """
    if not OPENPYXL_OK:
        print("\033[31m  openpyxl não instalado. Execute: pip install openpyxl\033[0m")
        return False

    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    _aba_resumo(wb, scores, score_geral, output_file)
    _aba_resultados(wb, resultados)
    _aba_plano(wb, resultados)

    if delta:
        _aba_comparacao(wb, delta)

    wb.save(output_file)
    return True